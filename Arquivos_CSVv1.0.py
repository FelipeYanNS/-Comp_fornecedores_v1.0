import unicodedata
import pandas as pd
from pandas.tseries.offsets import MonthEnd
import ipywidgets as widgets
import numpy as np
from google.colab import files
from google.colab import drive
import chardet
import string
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font

# Caminho da pasta no Google Drive
caminho_pasta = '/content/drive/MyDrive/Lançamentos Contábeis/'

# Função para remover acentos
def remover_acentos(txt):
    return unicodedata.normalize('NFKD', txt).encode('ASCII', 'ignore').decode('ASCII')

# Listar arquivos CSV na pasta
arquivos_csv = []
for raiz, dirs, arquivos in os.walk(caminho_pasta):
    for nome in arquivos:
        if nome.lower().endswith('.csv'):
            caminho_completo = os.path.join(raiz, nome)
            arquivos_csv.append(caminho_completo)

# Criar DataFrame com os arquivos e extrair o nome da empresa
df_arquivos = pd.DataFrame(arquivos_csv, columns=['Caminho Completo'])
df_arquivos['Arquivo CSV'] = df_arquivos['Caminho Completo'].apply(os.path.basename)
df_arquivos['Diretoria'] = df_arquivos['Arquivo CSV'].str.extract(r'^(.*?)\s*-\s*')

# Primeiras letras maiúsculas no seletor
empresas = sorted([
    string.capwords(str(e).strip().lower())
    for e in df_arquivos['Diretoria'].dropna().unique()
])

# Função para executada o botão de geração de arquivo (primeiro botão - verde)

def ao_clicar_botao1(botao=None):
    empresa_escolhida = seletor_empresa.value
    Intervalo = seletor_Intervalo.value
    print(f"\nEmpresa {empresa_escolhida}, no intervalo de {Intervalo}")

    arquivos_filtrados = []
    for raiz, dirs, arquivos in os.walk(caminho_pasta):
        for nome in arquivos:
            if nome.lower().endswith('.csv'):
                nome_sem_acentos = remover_acentos(nome).lower()
                empresa_normalizada = remover_acentos(empresa_escolhida).lower()
                if nome_sem_acentos.startswith(empresa_normalizada):
                    caminho_completo = os.path.join(raiz, nome)
                    arquivos_filtrados.append(caminho_completo)

    lista_dfs = []
    for caminho_arquivo in arquivos_filtrados:
        try:
            encoding = detectar_encoding(caminho_arquivo)
            df = pd.read_csv(caminho_arquivo, encoding=encoding, sep=None, engine='python')
            lista_dfs.append(df)
        except Exception as e:
            print(f"Erro ao ler o arquivo {caminho_arquivo}: {e}")

    if lista_dfs:
        # ------------------- 1 - PRÉ-PROCESSAMENTO DOS DADOS -------------------
        df1 = pd.concat(lista_dfs, ignore_index=True)
        df1['Valor'] = df1['Valor'].astype(str).str.replace(',', '.').str.strip()
        df1['Valor'] = pd.to_numeric(df1['Valor'])
        df1['Valor'] = df1['Valor']*1
        df1 = df1.replace(['', 'Documento', np.nan], '0')
        df1['Documento'] = pd.to_numeric(df1['Documento'], errors='coerce').where(
            pd.to_numeric(df1['Documento'], errors='coerce').notna(),
            df1['Documento']
        )
        df1['Conta'] = df1['Conta'].astype(str)
        df1 = df1[df1['Conta'].str.startswith('2.1.2.01')]

        df1_original = df1['Data'].copy()

        df1['Data'] = pd.to_datetime(df1['Data'], format='%d/%m/%Y', errors='coerce')
        df1['Ano'] = df1['Data'].dt.year

        ano_mais_recente = df1['Ano'].max()
        ano_mais_antigo = df1['Ano'].min()

        if Intervalo == "Todos os anos":
            ano_limite = ano_mais_antigo
        elif Intervalo == "Ano atual":
            ano_limite = ano_mais_recente
        elif Intervalo == "2 últimos anos":
            ano_limite = ano_mais_recente - 1
        elif Intervalo == "3 últimos anos":
            ano_limite = ano_mais_recente - 2
        else:
            print("Aviso: Valor de intervalo não reconhecido. Nenhum filtro de ano aplicado.")

        df1 = df1[df1['Ano'] >= ano_limite]
        primeiro_ano = ano_limite
        df1["Tipo"] = df1["Tipo"].astype(str)
        filtro_tipo = df1['Tipo'].str.startswith(('1', '0'))
        df1 = df1[(df1['Ano'] == primeiro_ano) | \
                          ((df1['Ano'] > primeiro_ano) & filtro_tipo)]

        df1['Data'] = df1_original.loc[df1.index]

        df1['Saldo'] = df1.apply(lambda row: row['Valor'] if row['Ação'] == 'C'
                                                                     else -1 * row['Valor'], axis=1)
        df1['Saldo'] = pd.to_numeric(df1['Saldo'], errors='coerce')
        soma_por_descricao = df1.groupby('Conta reduz.')['Saldo'].sum()
        descricoes_validas = soma_por_descricao[
            (soma_por_descricao.abs() > 0.009) &
            (soma_por_descricao.notnull())
        ].index
        df1_filtrado = df1[df1['Conta reduz.'].isin(descricoes_validas)]

        # ------------------- 2 - CONSTRUÇÃO DA TABELA DINÂMICA -----------------
        tabela_dinamica = pd.pivot_table(df1_filtrado,
                                         index=['Conta reduz.', 'Documento'],
                                         values='Saldo',
                                         aggfunc='sum',
                                         sort=False).reset_index()
        tabela_dinamica["Saldo"] = pd.to_numeric(tabela_dinamica["Saldo"], errors='coerce')

        # ------------------- 3 - COMPLEMENTAÇÃO COM METADADOS ------------------
        df_unico = df1.drop_duplicates(subset=["Conta reduz.", "Documento"], keep="first")[["Conta reduz.", "Documento", "Descrição.1", "Data"]]
        tabela_dinamica["Documento"] = tabela_dinamica["Documento"].astype(str)
        df_unico["Documento"] = df_unico["Documento"].astype(str)
        tabela_dinamica = tabela_dinamica.merge(
            df_unico,
            on=["Conta reduz.", "Documento"],
            how="left"
        )
        tabela_dinamica['Documento'] = pd.to_numeric(tabela_dinamica['Documento'], errors='coerce').where(
            pd.to_numeric(tabela_dinamica['Documento'], errors='coerce').notna(),
            tabela_dinamica['Documento']
        )
        tabela_dinamica = tabela_dinamica.rename(columns={
            "Descrição.1": "Fornecedor",
            "Documento": "Nota Fiscal",
            "Saldo": "Valores"
        })
        tabela_filtrada = tabela_dinamica[
            (tabela_dinamica["Valores"].abs() > 0.009) &
            (tabela_dinamica["Valores"].notnull())
        ]

        colunas_ordenadas = ["Conta reduz.", "Fornecedor", "Nota Fiscal", "Data", "Valores"]
        tabela_filtrada = tabela_filtrada[colunas_ordenadas]

        # ------------------- ETAPA 6: EXPORTAÇÃO E DOWNLOAD -------------------

        arquivo = f"{empresa_escolhida} - Composição dos fornecedores.xlsx"
        tabela_filtrada.to_excel(arquivo, index=False)

        from openpyxl import load_workbook
        wb = load_workbook(arquivo)
        ws = wb.active

        colunas_largura = {
            'A': 15,
            'B': 45,
            'C': 15,
            'D': 15,
            'E': 15
        }

        for col, largura in colunas_largura.items():
            ws.column_dimensions[col].width = largura

        wb.save(arquivo)
        display(files.download(arquivo))
    else:
        print(f"\nNão foi possível ler os arquivos da empresa '{empresa_escolhida}'.")

# Função para executada o botão de conferência (segundo botão - azul)

def ao_clicar_botao2(botao=None):
    empresa_escolhida = seletor_empresa.value
    Intervalo = seletor_Intervalo.value

    arquivos_filtrados = []
    for raiz, dirs, arquivos in os.walk(caminho_pasta):
        for nome in arquivos:
            if nome.lower().endswith('.csv'):
                nome_sem_acentos = remover_acentos(nome).lower()
                empresa_normalizada = remover_acentos(empresa_escolhida).lower()
                if nome_sem_acentos.startswith(empresa_normalizada):
                    caminho_completo = os.path.join(raiz, nome)
                    arquivos_filtrados.append(caminho_completo)

    lista_dfs = []
    for caminho_arquivo in arquivos_filtrados:
        try:
            encoding = detectar_encoding(caminho_arquivo)
            df = pd.read_csv(caminho_arquivo, encoding=encoding, sep=None, engine='python')
            lista_dfs.append(df)
        except Exception as e:
            print(f"Erro ao ler o arquivo {caminho_arquivo}: {e}")

    if lista_dfs:
        # ------------------- 1 - PRÉ-PROCESSAMENTO DOS DADOS -------------------
        df1 = pd.concat(lista_dfs, ignore_index=True)
        df1['Valor'] = df1['Valor'].astype(str).str.replace(',', '.').str.strip()
        df1['Valor'] = pd.to_numeric(df1['Valor'])
        df1['Valor'] = df1['Valor']*1
        df1 = df1.replace(['', 'Documento', np.nan], '0')
        df1['Documento'] = pd.to_numeric(df1['Documento'], errors='coerce').where(
            pd.to_numeric(df1['Documento'], errors='coerce').notna(),
            df1['Documento']
        )
        df1['Conta'] = df1['Conta'].astype(str)
        df1 = df1[df1['Conta'].str.startswith('2.1.2.01')]

        df1['Data'] = pd.to_datetime(df1['Data'], format='%d/%m/%Y', errors='coerce')
        df1['Ano'] = df1['Data'].dt.year

        ano_mais_recente = df1['Ano'].max()
        ano_mais_antigo = df1['Ano'].min()

        if Intervalo == "Todos os anos":
            ano_limite = ano_mais_antigo
        elif Intervalo == "Ano atual":
            ano_limite = ano_mais_recente
        elif Intervalo == "2 últimos anos":
            ano_limite = ano_mais_recente - 1
        elif Intervalo == "3 últimos anos":
            ano_limite = ano_mais_recente - 2
        else:
            print("Aviso: Valor de intervalo não reconhecido. Nenhum filtro de ano aplicado.")

        df1 = df1[df1['Ano'] >= ano_limite]
        primeiro_ano = ano_limite
        df1["Tipo"] = df1["Tipo"].astype(str)
        filtro_tipo = df1['Tipo'].str.startswith(('1', '0'))
        df1 = df1[(df1['Ano'] == primeiro_ano) | \
                          ((df1['Ano'] > primeiro_ano) & filtro_tipo)]

        df1['Saldo'] = df1.apply(lambda row: row['Valor'] if row['Ação'] == 'C'
                                                                     else -1 * row['Valor'], axis=1)
        Total_saldo = df1['Saldo'].sum()
        total_saldo1 = f"{Total_saldo:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

        display(print(f"\nEmpresa {empresa_escolhida}, no intervalo de {Intervalo}, possue um valor de R$ {total_saldo1}"))
    else:
        print(f"\nNão foi possível ler os arquivos da empresa '{empresa_escolhida}'.")


# Widgets

Intervalo = ['Ano atual', '2 últimos anos', '3 últimos anos', 'Todos os anos']

seletor_Intervalo = widgets.Dropdown(
    options=Intervalo,
    value='Todos os anos',
    description='Anos: ',
    layout=widgets.Layout(width='15%')
)

seletor_empresa = widgets.Dropdown(
    options=empresas,
    description='Empresa:',
    layout=widgets.Layout(width='25%')
)

botao_confirmar1 = widgets.Button(
    description='Baixar Arquivo',
    button_style='success',
    layout=widgets.Layout(width='150px')
)

botao_confirmar2 = widgets.Button(
    description='Conferência',
    button_style='info',
    layout=widgets.Layout(width='100px')
)

# Espaçador entre seletor e botões
espaco = widgets.Box(layout=widgets.Layout(width='25px'))

# Eventos dos botões
botao_confirmar1.on_click(ao_clicar_botao1)
botao_confirmar2.on_click(ao_clicar_botao2)

# Linha com seletor e botões + espaçador
linha_selecao = widgets.HBox(
    [seletor_Intervalo, seletor_empresa, espaco, botao_confirmar1, botao_confirmar2],
    layout=widgets.Layout(justify_content='flex-start', align_items='center', gap='10px')
)

# Caixa principal com título e borda
caixa_principal = widgets.VBox(
    [widgets.HTML(value="<h3>🔎 Análise e conciliação contábil</h3>"),
     linha_selecao],
    layout=widgets.Layout(
        border='solid 2px gray',
        padding='10px',
        width='80%',
        align_items='stretch'
    )
)

# Exibir interface
display(caixa_principal)
